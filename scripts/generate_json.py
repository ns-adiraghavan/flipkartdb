#!/usr/bin/env python3
"""
FK NextGen Lite dashboard — data generator.

Reads the "FK Dashboard Reports" Excel workbook and emits one JSON file per
dashboard tab into OUTPUT_DIR. Mirrors the TataCliq TAT pipeline: the SPA
fetches these JSON files at runtime, so a refresh = re-run this script and
replace the JSON (no rebuild).

Env vars:
  EXCEL_PATH  path to the .xlsx            (default: data/incoming/latest.xlsx)
  OUTPUT_DIR  where JSON files are written (default: public/data)

Source tabs used: Kam DUMP, BD DUMP, BD Internal Quality, Leads, Payment Tracking.
Helper tabs (Sort, Unique, Pivot*, Index) are ignored — raw dumps, counted as-is.
"""
import os, json, datetime
from collections import Counter, defaultdict
import openpyxl

EXCEL_PATH = os.environ.get("EXCEL_PATH", "data/incoming/latest.xlsx")
OUTPUT_DIR = os.environ.get("OUTPUT_DIR", "public/data")

# ---------- helpers ----------------------------------------------------------

def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s not in ("", "-", "#N/A", "NA", "N/A") else None
    return v

def title(s):
    return s.strip().title() if isinstance(s, str) else s

def load_sheet(wb, name):
    ws = wb[name]
    header = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(header) if h}
    return ws, header, idx

def rows_where(ws, key_col_1based):
    """Yield 1-based row numbers that have a non-empty key column (skips trailing blanks)."""
    for r in range(2, ws.max_row + 1):
        if norm(ws.cell(r, key_col_1based).value) is not None:
            yield r

def counter_to_list(counter, top=None):
    items = counter.most_common(top) if top else sorted(counter.items(), key=lambda x: -x[1])
    return [{"name": str(k), "value": v} for k, v in items]

def month_key(dt):
    return dt.strftime("%Y-%m")

# canonical maps for messy free-text -----------------------------------------
CONNECT_MAP = {"connected": "Connected", "not connected": "Not Connected"}
YESNO_MAP = {"yes": "Yes", "no": "No", "pending": "Pending"}

def canon(value, mapping, default_title=True):
    if value is None:
        return None
    key = str(value).strip().lower()
    if key in mapping:
        return mapping[key]
    return title(value) if default_title else value

# ---------- per-tab builders -------------------------------------------------

def build_leads(wb):
    ws, header, idx = load_sheet(wb, "Leads")
    c_brand = idx["Brand Name"] + 1
    c_type = idx["Lead Type"] + 1
    c_sid = idx["Seller ID"] + 1
    by_type = Counter()
    total = 0
    existing_fk = 0
    for r in rows_where(ws, c_brand):
        total += 1
        lt = norm(ws.cell(r, c_type).value)
        if lt:
            by_type[str(lt).strip()] += 1
        if norm(ws.cell(r, c_sid).value) is not None:
            existing_fk += 1

    # grouped source families
    grouped = Counter()
    for k, v in by_type.items():
        kl = k.lower()
        if "client inbound" in kl:
            grouped["Client Inbound"] += v
        elif "client lead" in kl:
            grouped["Client Lead Set"] += v
        elif "ns" in kl or "internal" in kl:
            grouped["NS Internal"] += v
        else:
            grouped["Other"] += v

    return {
        "kpis": {
            "total_leads": total,
            "distinct_sources": len(by_type),
            "existing_on_fk": existing_fk,
            "internal_leads": grouped.get("NS Internal", 0),
        },
        "by_type": counter_to_list(by_type),
        "by_group": counter_to_list(grouped),
    }

def build_bd(wb):
    ws, header, idx = load_sheet(wb, "BD DUMP")
    C = lambda n: idx[n] + 1
    c_date = C("Date of Calling")
    total = connected = 0
    by_bd = Counter(); conn_by_bd = defaultdict(int)
    connectivity = Counter(); disposition = Counter(); interest = Counter()
    d2c = Counter(); funnel = Counter(); mode = Counter(); lead_type = Counter()
    agreed_fk = Counter(); agreed_ng = Counter()
    by_day = Counter()
    agreed_fk_yes = 0; qualified = 0; interested = 0; callbacks = 0

    for r in rows_where(ws, c_date):
        total += 1
        bd = norm(ws.cell(r, C("Allocation")).value)
        if bd:
            by_bd[str(bd).strip()] += 1
        cs = canon(norm(ws.cell(r, C("Connectivity Status")).value), CONNECT_MAP)
        if cs:
            connectivity[cs] += 1
            if cs == "Connected":
                connected += 1
                if bd: conn_by_bd[str(bd).strip()] += 1
        disp = norm(ws.cell(r, C("Call Disposition Bucket")).value)
        if disp:
            disposition[str(disp).strip()] += 1
            if "call back" in str(disp).lower() or "callback" in str(disp).lower():
                callbacks += 1
        il = canon(norm(ws.cell(r, C("Interested Level")).value), {})
        if il and il.lower() != "yes":
            interest[il] += 1
            if il in ("Warm", "Hot"):
                interested += 1
        q = norm(ws.cell(r, C("D2C Qualification Status")).value)
        if q:
            d2c[title(q)] += 1
            if title(q) == "Qualified":
                qualified += 1
        fs = norm(ws.cell(r, C("Current Funnel Stage")).value)
        if fs: funnel[title(fs)] += 1
        md = norm(ws.cell(r, C("Mode of calling")).value)
        if md: mode[title(md)] += 1
        lt = norm(ws.cell(r, C("Lead Type")).value)
        if lt: lead_type[title(lt)] += 1
        af = canon(norm(ws.cell(r, C("Agreed On FK Onboarding")).value), YESNO_MAP)
        if af:
            agreed_fk[af] += 1
            if af == "Yes": agreed_fk_yes += 1
        an = canon(norm(ws.cell(r, C("Agreed for NextGen Services")).value), YESNO_MAP)
        if an: agreed_ng[an] += 1
        dv = ws.cell(r, c_date).value
        if isinstance(dv, datetime.datetime):
            by_day[dv.strftime("%Y-%m-%d")] += 1

    day_series = [{"date": d, "calls": n} for d, n in sorted(by_day.items())]
    conn_rate = round(connected / total * 100, 1) if total else 0
    # connected per BD stacked
    bd_list = [{"name": b, "calls": by_bd[b], "connected": conn_by_bd.get(b, 0)}
               for b, _ in by_bd.most_common()]

    return {
        "kpis": {
            "total_calls": total,
            "connected": connected,
            "connect_rate": conn_rate,
            "agents": len(by_bd),
            "agreed_onboarding": agreed_fk_yes,
            "d2c_qualified": qualified,
            "interested": interested,
            "callbacks": callbacks,
        },
        "by_bd": bd_list,
        "connectivity": counter_to_list(connectivity),
        "disposition": counter_to_list(disposition, top=12),
        "interest": counter_to_list(interest),
        "d2c": counter_to_list(d2c),
        "funnel_stage": counter_to_list(funnel),
        "mode": counter_to_list(mode),
        "lead_type": counter_to_list(lead_type),
        "agreed_fk": counter_to_list(agreed_fk),
        "by_day": day_series,
    }

def build_kam(wb):
    ws, header, idx = load_sheet(wb, "Kam DUMP")
    C = lambda n: idx[n] + 1
    c_date = C("DATE OF CALLING")
    total = connected = followups = 0
    by_kam = Counter(); health = Counter(); calltype = Counter()
    connectivity = Counter(); mode = Counter(); by_day = Counter()
    for r in rows_where(ws, c_date):
        total += 1
        k = norm(ws.cell(r, C("Assigned KAM")).value)
        if k: by_kam[str(k).strip()] += 1
        cs = canon(norm(ws.cell(r, C("Connectivity Status")).value), CONNECT_MAP)
        if cs:
            connectivity[cs] += 1
            if cs == "Connected": connected += 1
        h = norm(ws.cell(r, C("Health Status")).value)
        if h: health[title(h)] += 1
        ct = norm(ws.cell(r, C("Call Type")).value)
        if ct: calltype[title(ct)] += 1
        md = norm(ws.cell(r, C("Mode of calling")).value)
        if md: mode[title(md)] += 1
        fu = canon(norm(ws.cell(r, C("Follow-Up Required")).value), YESNO_MAP)
        if fu == "Yes": followups += 1
        dv = ws.cell(r, c_date).value
        if isinstance(dv, datetime.datetime):
            by_day[dv.strftime("%Y-%m-%d")] += 1
    return {
        "kpis": {
            "total_calls": total,
            "connected": connected,
            "connect_rate": round(connected / total * 100, 1) if total else 0,
            "kams": len(by_kam),
            "followups_required": followups,
            "at_risk": health.get("At Risk", 0),
        },
        "by_kam": counter_to_list(by_kam),
        "health": counter_to_list(health),
        "call_type": counter_to_list(calltype),
        "connectivity": counter_to_list(connectivity),
        "mode": counter_to_list(mode),
        "by_day": [{"date": d, "calls": n} for d, n in sorted(by_day.items())],
    }

def build_quality(wb):
    ws, header, idx = load_sheet(wb, "BD Internal Quality ")
    C = lambda n: idx[n] + 1
    c_score = C("Score")
    # the 10 SOP yes/no dimensions (long headers) — grab by keyword
    dim_keywords = {
        "SOP Adherence": "prescribed sales SOP",
        "Opening & Intro": "Opening & Introduction",
        "Discovery": "Seller Qualification & Discovery",
        "Benefits Pitch": "NextGen Lite Benefits Pitching",
        "Program Accuracy": "Program Details & Subscription",
        "Objection Handling": "Objection Handling & Value Selling",
        "Conversion": "Conversion & Commitment Taking",
        "Closing & Follow-up": "Closing & Follow-up",
    }
    dim_cols = {}
    for label, kw in dim_keywords.items():
        for h, i in idx.items():
            if kw.lower() in h.lower():
                dim_cols[label] = i + 1
                break

    scores = []; by_bd_scores = defaultdict(list); fatal = 0; audits = 0
    dim_pass = {k: 0 for k in dim_cols}; dim_total = {k: 0 for k in dim_cols}
    by_week = Counter()
    for r in rows_where(ws, c_score):
        sv = ws.cell(r, c_score).value
        if not isinstance(sv, (int, float)):
            continue
        audits += 1
        scores.append(sv)
        bd = norm(ws.cell(r, C("Email Address")).value)
        if bd: by_bd_scores[str(bd).strip()].append(sv)
        fe = canon(norm(ws.cell(r, C("Was there any Fatal error ?")).value), YESNO_MAP)
        if fe == "Yes": fatal += 1
        wk = norm(ws.cell(r, C("Week")).value)
        if wk: by_week[str(wk).strip()] += 1
        for label, col in dim_cols.items():
            val = norm(ws.cell(r, col).value)
            if val is not None:
                dim_total[label] += 1
                if str(val).strip().lower() in ("yes", "y"):
                    dim_pass[label] += 1

    avg = round(sum(scores) / len(scores), 1) if scores else 0
    passing = sum(1 for s in scores if s >= 60)
    bd_avg = [{"name": b, "value": round(sum(v) / len(v), 1), "audits": len(v)}
              for b, v in sorted(by_bd_scores.items(), key=lambda x: -sum(x[1]) / len(x[1]))]
    dim_rates = [{"name": k, "value": round(dim_pass[k] / dim_total[k] * 100, 1) if dim_total[k] else 0}
                 for k in dim_cols]
    # score distribution buckets
    buckets = Counter()
    for s in scores:
        b = "0-40" if s < 40 else "40-60" if s < 60 else "60-80" if s < 80 else "80-100"
        buckets[b] += 1
    dist = [{"name": b, "value": buckets.get(b, 0)} for b in ["0-40", "40-60", "60-80", "80-100"]]

    return {
        "kpis": {
            "audits": audits,
            "avg_score": avg,
            "pass_rate": round(passing / audits * 100, 1) if audits else 0,
            "fatal_errors": fatal,
            "bds_audited": len(by_bd_scores),
        },
        "by_bd": bd_avg,
        "sop_dimensions": dim_rates,
        "score_distribution": dist,
        "by_week": counter_to_list(by_week),
    }

def build_payment(wb):
    ws, header, idx = load_sheet(wb, "Payment Tracking")
    C = lambda n: idx[n] + 1
    c_sid = C("Seller ID")
    total = 0; rev_incl = 0.0; rev_excl = 0.0
    by_re = Counter(); rev_by_re = defaultdict(float)
    by_month = Counter(); by_program = Counter(); by_paytype = Counter()
    by_day_rev = defaultdict(float)
    tickets = []
    for r in rows_where(ws, c_sid):
        total += 1
        incl = ws.cell(r, C("Final Amount Including GST")).value
        excl = ws.cell(r, C("Amount Without GST")).value
        if isinstance(incl, (int, float)):
            rev_incl += incl; tickets.append(incl)
        if isinstance(excl, (int, float)):
            rev_excl += excl
        re_name = norm(ws.cell(r, C("RE NAME")).value)
        if re_name:
            key = str(re_name).strip().title()
            by_re[key] += 1
            if isinstance(incl, (int, float)):
                rev_by_re[key] += incl
        mn = norm(ws.cell(r, C("Month Name")).value)
        if mn: by_month[str(mn).strip()] += 1
        pg = norm(ws.cell(r, C("Program Name")).value)
        if pg: by_program[str(pg).strip()] += 1
        pt = norm(ws.cell(r, C("Payment Type")).value)
        if pt: by_paytype[str(pt).strip()] += 1
        dv = ws.cell(r, C("Date (Payment)")).value
        if isinstance(dv, datetime.datetime) and isinstance(incl, (int, float)):
            by_day_rev[dv.strftime("%Y-%m-%d")] += incl

    return {
        "kpis": {
            "sellers_paid": total,
            "revenue_incl_gst": round(rev_incl),
            "revenue_excl_gst": round(rev_excl),
            "gst_collected": round(rev_incl - rev_excl),
            "avg_ticket": round(rev_incl / len(tickets)) if tickets else 0,
        },
        "revenue_by_re": [{"name": k, "value": round(v), "sellers": by_re[k]}
                          for k, v in sorted(rev_by_re.items(), key=lambda x: -x[1])],
        "sellers_by_re": counter_to_list(by_re),
        "by_month": counter_to_list(by_month),
        "by_program": counter_to_list(by_program),
        "by_paytype": counter_to_list(by_paytype),
        "revenue_by_day": [{"date": d, "revenue": round(v)} for d, v in sorted(by_day_rev.items())],
    }

def build_overview(leads, bd, kam, quality, payment):
    funnel = [
        {"stage": "Leads Sourced", "value": leads["kpis"]["total_leads"]},
        {"stage": "Calls Made", "value": bd["kpis"]["total_calls"]},
        {"stage": "Connected", "value": bd["kpis"]["connected"]},
        {"stage": "Interested", "value": bd["kpis"]["interested"]},
        {"stage": "Agreed Onboarding", "value": bd["kpis"]["agreed_onboarding"]},
        {"stage": "Paid / Onboarded", "value": payment["kpis"]["sellers_paid"]},
    ]
    return {
        "kpis": {
            "total_leads": leads["kpis"]["total_leads"],
            "total_calls": bd["kpis"]["total_calls"],
            "connect_rate": bd["kpis"]["connect_rate"],
            "sellers_paid": payment["kpis"]["sellers_paid"],
            "revenue_incl_gst": payment["kpis"]["revenue_incl_gst"],
            "avg_qa_score": quality["kpis"]["avg_score"],
            "kam_calls": kam["kpis"]["total_calls"],
            "agents": bd["kpis"]["agents"],
        },
        "funnel": funnel,
        "calls_by_day": bd["by_day"],
        "revenue_by_re": payment["revenue_by_re"],
        "leads_by_group": leads["by_group"],
    }

# ---------- main -------------------------------------------------------------

def main():
    print(f"Reading {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    leads = build_leads(wb)
    bd = build_bd(wb)
    kam = build_kam(wb)
    quality = build_quality(wb)
    payment = build_payment(wb)
    overview = build_overview(leads, bd, kam, quality, payment)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    meta = {"generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "source": os.path.basename(EXCEL_PATH)}
    outputs = {"overview": overview, "leads": leads, "bd": bd,
               "kam": kam, "quality": quality, "payment": payment, "meta": meta}
    for name, data in outputs.items():
        path = os.path.join(OUTPUT_DIR, f"{name}.json")
        with open(path, "w") as f:
            json.dump(data, f, indent=2, default=str)
        print(f"  wrote {path}")
    print("Done.")

if __name__ == "__main__":
    main()
